#!/usr/bin/env python3
"""Build the INTEGRI information-gathering workbook."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT = "/home/user/SECUMAX-SECURITY-WEBSITE/handover/INTEGRI-Information-Request.xlsx"

# ---------- palette ----------
RED      = "C00018"
BLACK    = "111111"
HEAD_BG  = "111111"
BAND_BG  = "F2F2F2"
FILL_IN  = "FFF2CC"   # cells the client edits
FILL_HD  = "FDECEC"
GREY     = "666666"

F = "Arial"
h1   = Font(name=F, size=16, bold=True, color="FFFFFF")
h2   = Font(name=F, size=12, bold=True, color="FFFFFF")
sect = Font(name=F, size=11, bold=True, color=RED)
bold = Font(name=F, size=10, bold=True)
base = Font(name=F, size=10)
smal = Font(name=F, size=9, color=GREY)
ital = Font(name=F, size=9, italic=True, color=GREY)

thin = Side(style="thin", color="D9D9D9")
box  = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical="top")
wrapc= Alignment(wrap_text=True, vertical="center")
ctr  = Alignment(horizontal="center", vertical="center")

def fill(hexcode): return PatternFill("solid", fgColor=hexcode)

# ============================================================
# CONTENT
# ============================================================
SERVICES = [
("01", "Investigation Services", [
 ("Corporate & internal investigations",
  "We can investigate internal misconduct, theft or policy breaches inside a client's business and deliver a written report."),
 ("Fraud & financial crime enquiry",
  "We have someone able to investigate fraud, establish how money moved, and document it to an evidentiary standard."),
 ("Surveillance & counter-surveillance",
  "We have trained operatives and equipment to run lawful physical surveillance, and to detect surveillance run against a client."),
 ("Background screening & vetting",
  "We can run pre-employment and due-diligence background checks, with the subject's consent where the law requires it."),
 ("Insurance claim investigation",
  "We take instructions from insurers or brokers to investigate suspect claims and report findings."),
 ("Asset tracing & due diligence",
  "We can trace assets, company interests and directorships, and compile a due-diligence report."),
 ("Missing persons tracing",
  "We take on missing-person and skip-tracing work."),
 ("Matrimonial & domestic enquiry",
  "We take on private matrimonial and domestic investigations for individual clients."),
]),
("02", "Forensic Services", [
 ("Crime scene processing & documentation",
  "We attend scenes and photograph, document, and recover physical evidence using proper method."),
 ("Fingerprint collection & comparison",
  "We can BOTH lift latent prints AND have someone competent to compare and identify them. Lifting alone is not enough for this line."),
 ("Questioned document examination",
  "We have a trained document examiner who can examine signatures, handwriting and alterations."),
 ("Digital & mobile device forensics",
  "We have forensic acquisition tools and a competent examiner who can image and analyse devices defensibly."),
 ("Forensic auditing",
  "We have someone with an accounting or audit qualification able to conduct a forensic audit."),
 ("Evidence handling & chain of custody",
  "We operate a documented chain-of-custody procedure: exhibit register, sealed and labelled packaging, signed handovers."),
 ("Expert witness testimony",
  "We have someone who could realistically be accepted by a court as an expert in their field and testify."),
 ("Court-ready forensic reporting",
  "We produce reports written to a standard suitable for court, a CCMA matter or a disciplinary hearing."),
]),
("03", "Polygraph Services", [
 ("Pre-employment screening examinations",
  "We have a trained polygraph examiner and working instrumentation available for pre-employment screening."),
 ("Periodic integrity testing",
  "We can run scheduled, repeat integrity examinations for a client's staff."),
 ("Specific-issue (incident) examinations",
  "We run single-issue examinations tied to a specific incident."),
 ("Theft & shrinkage investigations",
  "We use polygraph examinations as one input in theft and stock-loss investigations."),
 ("Disciplinary hearing support",
  "We provide examination reports for use in an employer's disciplinary process, as supporting material rather than proof."),
 ("Internal integrity programmes",
  "We can design and run an ongoing integrity-testing programme for a client."),
 ("Structured pre-test interviewing",
  "Every examination includes a structured pre-test interview and written informed consent."),
 ("Documented examination reporting",
  "Every examination produces a written report recording the questions, the data and the examiner's opinion."),
]),
("04", "Security Services", [
 ("Security risk assessments",
  "We can survey a site or an operation and produce a written risk assessment with recommendations."),
 ("Access control & visitor management",
  "We can design and implement access-control and visitor-management systems."),
 ("CCTV design, installation & monitoring",
  "We can specify, physically install and monitor CCTV. If we only advise or only monitor, mark this FALSE and say which."),
 ("Alarm systems & response coordination",
  "We can install or specify alarm systems and coordinate an armed-response arrangement."),
 ("Perimeter & physical security design",
  "We can design perimeter and physical security measures — fencing, lighting, barriers, layout."),
 ("Control room & monitoring solutions",
  "We can set up or supply a control room / monitoring capability."),
 ("Security policy & SOP development",
  "We write security policies, post orders and standard operating procedures for clients."),
 ("Security audits & compliance reviews",
  "We audit an existing security setup against its own policies or applicable requirements."),
]),
("05", "Protection Services", [
 ("Close protection officers",
  "We have PSIRA-registered close protection officers, graded for the role, available for deployment."),
 ("Executive & corporate protection",
  "We provide protection details for company executives and corporate clients."),
 ("Family & residential protection",
  "We provide protection covering a principal's family and residence."),
 ("Secure chauffeur & transport",
  "We can provide trained security drivers and secure transport."),
 ("Travel risk management & advance work",
  "We conduct route planning, venue advance work and travel risk assessment before a movement."),
 ("Event & dignitary protection",
  "We provide protection at events and for visiting dignitaries."),
 ("Threat & vulnerability assessment",
  "We assess threats against an individual and produce a written protective assessment."),
 ("Female close protection officers",
  "We have at least one female close protection officer on the team or reliably available."),
]),
("06", "Guarding Services", [
 ("Static & site guarding",
  "We deploy PSIRA-registered guards to static posts and sites."),
 ("Access control officers",
  "We supply officers specifically for gate and access-control duty."),
 ("Residential estate guarding",
  "We guard residential estates and complexes."),
 ("Retail & commercial guarding",
  "We guard retail and commercial premises."),
 ("Construction & industrial site security",
  "We secure construction sites and industrial premises."),
 ("Mobile patrols & guard monitoring",
  "We run mobile patrols and monitor guards on shift, with a record of visits or checkpoints."),
 ("Reaction & response officers",
  "We field reaction or response officers, whether our own or through a formal arrangement."),
 ("Supervision & control room staffing",
  "We supply supervisors and control-room operators, not only guards."),
]),
("07", "Specialized Services", [
 ("Accredited firearm training",
  "We hold current accreditation to deliver firearm competency training, or deliver it through an accredited partner."),
 ("Tactical & refresher training",
  "We run tactical and refresher shooting or handling courses beyond basic competency."),
 ("Valuable goods & cash escorts",
  "We escort valuable goods or cash. Note: cash-in-transit is separately regulated — confirm we are permitted for what we advertise."),
 ("Asset & vehicle recovery",
  "We trace and recover stolen assets and vehicles, lawfully and with SAPS involvement where required."),
 ("Loss prevention programmes",
  "We design and run ongoing loss-prevention programmes for clients."),
 ("Undercover operative placement",
  "We place undercover operatives inside a client's workplace, with a lawful basis and the client's written mandate."),
 ("Crowd & labour unrest support",
  "We provide lawful, defensive, de-escalation-first support to a client's site during unrest. This is NOT strike-breaking."),
 ("Bespoke risk solutions",
  "We take on bespoke briefs that fall outside the standard categories."),
]),
]

# ------------------------------------------------------------------
# Status against the client's Business Profile (see SERVICE-STATUS.md).
# HOLD services are parked on the site and are what this workbook asks about.
# ------------------------------------------------------------------
GROUP = {
 "A": "Significant capability absent from your profile",
 "B": "Needs a named, qualified person",
 "C": "Plausible extension of confirmed work",
 "D": "Superseded or narrow",
}
HOLD = {
 # Group A
 "CCTV design, installation & monitoring":"A", "Alarm systems & response coordination":"A",
 "Access control & visitor management":"A", "Control room & monitoring solutions":"A",
 "Perimeter & physical security design":"A", "Static & site guarding":"A",
 "Residential estate guarding":"A", "Retail & commercial guarding":"A",
 "Construction & industrial site security":"A", "Mobile patrols & guard monitoring":"A",
 "Reaction & response officers":"A", "Corporate & internal investigations":"A",
 # Group B
 "Fingerprint collection & comparison":"B", "Questioned document examination":"B",
 "Digital & mobile device forensics":"B", "Forensic auditing":"B",
 "Expert witness testimony":"B", "Crime scene processing & documentation":"B",
 "Evidence handling & chain of custody":"B", "Court-ready forensic reporting":"B",
 "Fraud & financial crime enquiry":"B", "Surveillance & counter-surveillance":"B",
 # Group C
 "Asset tracing & due diligence":"C", "Insurance claim investigation":"C",
 "Matrimonial & domestic enquiry":"C", "Periodic integrity testing":"C",
 "Theft & shrinkage investigations":"C", "Internal integrity programmes":"C",
 "Security policy & SOP development":"C", "Security audits & compliance reviews":"C",
 "Family & residential protection":"C", "Travel risk management & advance work":"C",
 "Threat & vulnerability assessment":"C", "Access control officers":"C",
 # Group D
 "Missing persons tracing":"D", "Valuable goods & cash escorts":"D",
 "Tactical & refresher training":"D", "Structured pre-test interviewing":"D",
 "Documented examination reporting":"D", "Supervision & control room staffing":"D",
 "Female close protection officers":"D", "Asset & vehicle recovery":"D",
 "Loss prevention programmes":"D", "Undercover operative placement":"D",
}
LIVE = [
 ("01 Investigation", ["Track and Trace","Vetting","Criminal Record Checks","Extortion Cases",
                       "Evictions","Kidnap and Ransom Cases","Polygraphs"]),
 ("02 Forensic",      ["No itemised list — scope defined per assignment, per your profile"]),
 ("03 Polygraph",     ["Examinations supporting investigations","Screening examinations",
                       "Internal enquiry support"]),
 ("04 Security",      ["Risk-based security support","Asset Protection","Bullion Runs",
                       "High Value Assets in Transit"]),
 ("05 Protection",    ["Executive Close Protection","Corporate Executive Close Protection",
                       "Special Event Security","Secure Drivers"]),
 ("06 Guarding",      ["No itemised list — scope defined per assignment, per your profile"]),
 ("07 Specialized",   ["Mining Security","Illegal Mining Prevention Teams",
                       "Riot / Civil Unrest Control","Dedicated Searches","Mining Investigations",
                       "Firearm Training","Corporate Training","Riot Control Training",
                       "Security Training","Bespoke risk solutions"]),
]

COMPANY = [
 ("Registered company name", "INTEGRI Forensic Services (Pty) Ltd", "Confirmed from CIPC certificate", False),
 ("Company registration number", "2026/561988/07", "Confirmed from CIPC certificate", False),
 ("Trading name used on the website", "INTEGRI Forensic and Protection Services", "Confirmed", False),
 ("PSIRA registration number (business)", "", "Needed — attach the certificate", True),
 ("PSIRA registration expiry / renewal date", "", "Needed", True),
 ("VAT number (if registered)", "", "Leave blank if not VAT registered", True),
 ("Firearm training accreditation body", "", "e.g. SAPS, SASSETA", True),
 ("Firearm training accreditation number", "", "Needed — attach the certificate", True),
 ("Firearm competency types covered", "", "e.g. handgun, shotgun, rifle, self-defence, business purposes", True),
 ("B-BBEE level", "", "Leave blank if not yet assessed", True),
 ("Business address to publish on the site", "", "The CIPC address is residential and is NOT published. Give a commercial or PO Box address, or write NONE.", True),
 ("Office hours (admin and quotes)", "", "e.g. Mon-Fri 08:00-17:00", True),
 ("After-hours / emergency availability", "", "Is the 24/7 line genuinely answered at any hour?", True),
 ("Provinces actually serviced", "", "The site currently says 'national'. List the real provinces.", True),
 ("Website domain registered to INTEGRI?", "", "Yes / No", True),
 ("Is ops@ ... .com live and monitored?", "", "Yes / No", True),
 ("LinkedIn URL", "", "Leave blank if none", True),
 ("Facebook URL", "", "Leave blank if none", True),
 ("Instagram URL", "", "Leave blank if none", True),
 ("Professional indemnity insurer & cover", "", "Not published — used when clients ask", True),
 ("Public liability insurer & cover", "", "Not published — used when clients ask", True),
]

PEOPLE = [
 ("Etienne", "Full legal name", "Job title to publish", "Phone: +27 71 118 3257 (confirmed)",
  "The site currently says 'Director'. CIPC lists only Anri Coetser as a director — give the correct title."),
 ("Jacques", "Full legal name", "Job title to publish", "Phone: +27 67 161 2570 (confirmed)",
  "The site currently says 'Director'. Give the correct title."),
 ("Anri Coetser", "Full legal name", "Job title to publish", "Phone",
  "Sole CIPC-registered director. Should she appear on the website at all? Yes / No."),
 ("Information Officer", "Full legal name", "Position", "Email",
  "Required by POPIA because the website collects enquiries. Usually a director. Must be registered with the Information Regulator."),
 ("Deputy Information Officer", "Full legal name", "Position", "Email", "Optional."),
 ("Polygraph examiner", "Full name", "Qualification / training body", "Association membership",
  "Backs the polygraph pages."),
 ("Firearm instructor", "Full name", "Accreditation number", "Contact",
  "Backs the firearm training claim."),
 ("Forensic expert witness", "Full name", "Field of expertise", "Qualification",
  "The site offers expert witness testimony — who would actually testify?"),
]

DOCUMENTS = [
 ("PSIRA registration certificate — business", "Publishes the registration number and supports the compliance statement on every page"),
 ("PSIRA certificates — individual officers", "The guarding and protection pages state officers are registered and graded"),
 ("Firearm training accreditation certificate", "Supports the accredited firearm training claim"),
 ("Polygraph examiner qualification(s)", "Supports the polygraph division"),
 ("Tax clearance certificate", "Corporate clients and tenders"),
 ("VAT registration certificate", "Only if VAT registered — for the footer"),
 ("B-BBEE certificate or sworn affidavit", "Tenders and corporate procurement"),
 ("Professional indemnity insurance schedule", "Clients ask before appointing"),
 ("Public liability insurance schedule", "Clients ask before appointing"),
 ("Company letterhead / brand guide", "Keeps documents consistent with the site"),
 ("Logo — full lockup, transparent PNG", "Crest received. Lockup with wordmark still wanted for documents."),
 ("Team photographs (optional)", "Would replace the initial-letter avatars on the site"),
 ("CIPC registration certificate", "RECEIVED — no action needed"),
 ("Logo — crest, transparent PNG", "RECEIVED — now live on the site"),
]

# ============================================================
# BUILD
# ============================================================
wb = Workbook()

def banner(ws, title, subtitle, width):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    c = ws.cell(1, 1, title); c.font = h1; c.fill = fill(HEAD_BG); c.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[1].height = 34
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width)
    c = ws.cell(2, 1, subtitle); c.font = Font(name=F, size=9, color="FFFFFF"); c.fill = fill(RED)
    c.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[2].height = 20

# ---------------- Sheet 1: Start here ----------------
ws = wb.active; ws.title = "Start here"
banner(ws, "INTEGRI — Information Request",
       "Everything we need to finish the website. Fill in the shaded cells and send this file back.", 4)
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 34
ws.column_dimensions["C"].width = 62
ws.column_dimensions["D"].width = 34

rows = [
 ("", "", ""),
 ("HOW TO USE THIS FILE", "", ""),
 ("Tab 2 — Services", "56 services currently on the website. For each one, choose TRUE or FALSE. "
  "TRUE means INTEGRI can deliver it today. Anything marked FALSE is removed from the website.", ""),
 ("Tab 3 — Company details", "Fill in the shaded cells. Leave anything blank that does not apply.", ""),
 ("Tab 4 — People", "Names and correct job titles. The Information Officer is required by law.", ""),
 ("Tab 5 — Documents", "Tick off the certificates as you attach them to your reply.", ""),
 ("", "", ""),
 ("WHAT THE SHADING MEANS", "", ""),
 ("Shaded cells", "You fill these in.", ""),
 ("Unshaded cells", "Already confirmed — please just check they are correct.", ""),
 ("", "", ""),
 ("EXAMPLE — how a Services row should look when completed", "", ""),
]
r = 4
for a, b, c in rows:
    ws.cell(r, 2, a).font = sect if a.isupper() and a else bold
    cell = ws.cell(r, 3, b); cell.font = base; cell.alignment = wrap
    r += 1

# example row
hdr = ["Service", "What has to be true", "TRUE / FALSE", "Your note"]
for i, t in enumerate(hdr):
    c = ws.cell(r, 2 + i, t)
    c.font = Font(name=F, size=9, bold=True, color="FFFFFF"); c.fill = fill(HEAD_BG)
    c.alignment = wrapc; c.border = box
r += 1
ex = ["Fingerprint collection & comparison",
      "We can BOTH lift latent prints AND compare them.",
      "FALSE",
      "We can lift prints but we send comparison to an external expert — please reword."]
for i, v in enumerate(ex):
    c = ws.cell(r, 2 + i, v); c.font = ital; c.alignment = wrap; c.border = box
    if i >= 2: c.fill = fill(FILL_IN)
ws.row_dimensions[r].height = 40
r += 2
c = ws.cell(r, 2, "Questions? Reply to this email and we will walk through it with you.")
c.font = smal

# ---------------- Sheet 2: Services to verify ----------------
ws = wb.create_sheet("Services to verify")
banner(ws, "44 services we have taken OFF your website",
       "These are not in your business profile, so we parked them. Mark each one TRUE if INTEGRI "
       "does it, and we will put it back.", 7)
for i, w in enumerate([6, 24, 32, 46, 26, 13, 34], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

HROW = 4
heads = ["Ref", "Division", "Service", "What has to be true for us to advertise this",
         "Why we parked it", "TRUE / FALSE", "Your note"]
for i, t in enumerate(heads, start=1):
    c = ws.cell(HROW, i, t)
    c.font = Font(name=F, size=9, bold=True, color="FFFFFF"); c.fill = fill(HEAD_BG)
    c.alignment = wrapc; c.border = box
ws.row_dimensions[HROW].height = 30
ws.freeze_panes = "A5"

dv = DataValidation(type="list", formula1='"TRUE,FALSE,NEEDS REWORDING"', allow_blank=True, showDropDown=False)
dv.error = "Choose TRUE, FALSE or NEEDS REWORDING from the list."
dv.errorTitle = "Pick one of the options"
ws.add_data_validation(dv)

r = HROW + 1
first_data = r
n_hold = 0
for no, div, items in SERVICES:
    held = [(n, t) for n, t in items if n in HOLD]
    if not held:
        continue
    dr = ws.cell(r, 1, no); dr.font = Font(name=F, size=10, bold=True, color=RED)
    dr.fill = fill(BAND_BG); dr.alignment = ctr; dr.border = box
    dc = ws.cell(r, 2, div); dc.font = Font(name=F, size=10, bold=True)
    dc.fill = fill(BAND_BG); dc.alignment = wrapc; dc.border = box
    for col in range(3, 8):
        cc = ws.cell(r, col); cc.fill = fill(BAND_BG); cc.border = box
    ws.cell(r, 3, f"{len(held)} parked").font = smal
    ws.row_dimensions[r].height = 20
    r += 1
    for i, (name, truth) in enumerate(held, start=1):
        n_hold += 1
        g = HOLD[name]
        ws.cell(r, 1, f"{no}.{i}").font = smal
        ws.cell(r, 1).alignment = ctr; ws.cell(r, 1).border = box
        ws.cell(r, 2).border = box
        c = ws.cell(r, 3, name); c.font = bold; c.alignment = wrap; c.border = box
        c = ws.cell(r, 4, truth); c.font = base; c.alignment = wrap; c.border = box
        c = ws.cell(r, 5, f"{g} — {GROUP[g]}"); c.font = smal; c.alignment = wrap; c.border = box
        c = ws.cell(r, 6); c.fill = fill(FILL_IN); c.alignment = ctr; c.border = box
        c.font = Font(name=F, size=10, bold=True); dv.add(c)
        c = ws.cell(r, 7); c.fill = fill(FILL_IN); c.alignment = wrap; c.border = box; c.font = base
        ws.row_dimensions[r].height = 34
        r += 1
last_data = r - 1

r += 1
ws.cell(r, 3, "Progress").font = sect
ws.cell(r, 4, "Answered").font = bold
ws.cell(r, 6, f'=COUNTIF(F{first_data}:F{last_data},"TRUE")+COUNTIF(F{first_data}:F{last_data},"FALSE")'
              f'+COUNTIF(F{first_data}:F{last_data},"NEEDS REWORDING")').font = bold
ws.cell(r, 7, f"of {n_hold} parked services").font = smal
r += 1
ws.cell(r, 4, "TRUE — we will put these back").font = base
ws.cell(r, 6, f'=COUNTIF(F{first_data}:F{last_data},"TRUE")').font = base
r += 1
ws.cell(r, 4, "FALSE — these stay off").font = base
ws.cell(r, 6, f'=COUNTIF(F{first_data}:F{last_data},"FALSE")').font = base

# ---------------- Sheet 3: Services now live ----------------
ws = wb.create_sheet("Services now live")
banner(ws, "What your website says today",
       "Taken straight from your business profile. Please check nothing here is wrong.", 3)
for col, w in zip("ABC", (26, 52, 20)):
    ws.column_dimensions[col].width = w
HROW = 4
for i, t in enumerate(["Division", "Service", "Correct?"], start=1):
    c = ws.cell(HROW, i, t); c.font = Font(name=F, size=9, bold=True, color="FFFFFF")
    c.fill = fill(HEAD_BG); c.alignment = wrapc; c.border = box
ws.row_dimensions[HROW].height = 24
ws.freeze_panes = "A5"

dv3 = DataValidation(type="list", formula1='"YES,NO"', allow_blank=True, showDropDown=False)
ws.add_data_validation(dv3)

r = HROW + 1
for div, items in LIVE:
    c = ws.cell(r, 1, div); c.font = Font(name=F, size=10, bold=True, color=RED)
    c.fill = fill(BAND_BG); c.alignment = wrapc; c.border = box
    for col in (2, 3):
        cc = ws.cell(r, col); cc.fill = fill(BAND_BG); cc.border = box
    ws.cell(r, 2, f"{len(items)} live").font = smal
    r += 1
    for name in items:
        ws.cell(r, 1).border = box
        c = ws.cell(r, 2, name); c.font = base; c.alignment = wrap; c.border = box
        c = ws.cell(r, 3); c.fill = fill(FILL_IN); c.alignment = ctr; c.border = box
        c.font = Font(name=F, size=10, bold=True); dv3.add(c)
        ws.row_dimensions[r].height = 22
        r += 1

# ---------------- Sheet 3: Company details ----------------
ws = wb.create_sheet("Company details")
banner(ws, "Company details", "Fill in the shaded cells. Leave blank anything that does not apply yet.", 3)
for col, w in zip("ABC", (44, 40, 58)):
    ws.column_dimensions[col].width = w
HROW = 4
for i, t in enumerate(["Detail", "Your answer", "Notes"], start=1):
    c = ws.cell(HROW, i, t); c.font = Font(name=F, size=9, bold=True, color="FFFFFF")
    c.fill = fill(HEAD_BG); c.alignment = wrapc; c.border = box
ws.row_dimensions[HROW].height = 24
ws.freeze_panes = "A5"
r = HROW + 1
for label, value, note, editable in COMPANY:
    c = ws.cell(r, 1, label); c.font = bold; c.alignment = wrap; c.border = box
    c = ws.cell(r, 2, value); c.font = base; c.alignment = wrap; c.border = box
    if editable: c.fill = fill(FILL_IN)
    c = ws.cell(r, 3, note); c.font = smal; c.alignment = wrap; c.border = box
    ws.row_dimensions[r].height = 30
    r += 1

# ---------------- Sheet 4: People ----------------
ws = wb.create_sheet("People")
banner(ws, "People and roles", "Names and the job title we should publish. The Information Officer is required by POPIA.", 5)
for col, w in zip("ABCDE", (26, 32, 30, 30, 56)):
    ws.column_dimensions[col].width = w
HROW = 4
for i, t in enumerate(["Person / role", "Full legal name", "Job title to publish", "Contact", "Why we need this"], start=1):
    c = ws.cell(HROW, i, t); c.font = Font(name=F, size=9, bold=True, color="FFFFFF")
    c.fill = fill(HEAD_BG); c.alignment = wrapc; c.border = box
ws.row_dimensions[HROW].height = 24
ws.freeze_panes = "A5"
r = HROW + 1
for who, f1, f2, f3, note in PEOPLE:
    c = ws.cell(r, 1, who); c.font = bold; c.alignment = wrap; c.border = box
    for col in (2, 3, 4):
        c = ws.cell(r, col); c.fill = fill(FILL_IN); c.border = box
        c.alignment = wrap; c.font = base
    c = ws.cell(r, 5, note); c.font = smal; c.alignment = wrap; c.border = box
    ws.row_dimensions[r].height = 44
    r += 1
r += 1
ws.cell(r, 1, "Do NOT include ID numbers or home addresses — we do not publish them and do not need them.").font = Font(name=F, size=9, bold=True, color=RED)

# ---------------- Sheet 5: Documents ----------------
ws = wb.create_sheet("Documents")
banner(ws, "Documents to attach", "Tick each one off as you attach it to your reply.", 4)
for col, w in zip("ABCD", (6, 52, 62, 20)):
    ws.column_dimensions[col].width = w
HROW = 4
for i, t in enumerate(["#", "Document", "Why we need it", "Attached?"], start=1):
    c = ws.cell(HROW, i, t); c.font = Font(name=F, size=9, bold=True, color="FFFFFF")
    c.fill = fill(HEAD_BG); c.alignment = wrapc; c.border = box
ws.row_dimensions[HROW].height = 24
ws.freeze_panes = "A5"

dv2 = DataValidation(type="list", formula1='"YES,NOT YET,NOT APPLICABLE"', allow_blank=True, showDropDown=False)
ws.add_data_validation(dv2)

r = HROW + 1
dfirst = r
for i, (doc, why) in enumerate(DOCUMENTS, start=1):
    received = why.startswith("RECEIVED")
    c = ws.cell(r, 1, i); c.font = smal; c.alignment = ctr; c.border = box
    c = ws.cell(r, 2, doc); c.font = bold; c.alignment = wrap; c.border = box
    c = ws.cell(r, 3, why); c.font = smal; c.alignment = wrap; c.border = box
    c = ws.cell(r, 4); c.border = box; c.alignment = ctr; c.font = Font(name=F, size=10, bold=True)
    if received:
        c.value = "YES"
        c.font = Font(name=F, size=10, bold=True, color="1F7A4D")
    else:
        c.fill = fill(FILL_IN); dv2.add(c)
    ws.row_dimensions[r].height = 30
    r += 1
dlast = r - 1
r += 1
ws.cell(r, 2, "Attached so far").font = sect
ws.cell(r, 4, f'=COUNTIF(D{dfirst}:D{dlast},"YES")').font = bold
ws.cell(r, 3, f"of {len(DOCUMENTS)} documents").font = smal

import os
os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)
print("written:", OUT)
